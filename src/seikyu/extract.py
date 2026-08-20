"""発注書 → PurchaseOrder の構造化抽出（Claude API）。

PDF は document ブロック、画像は image ブロック、Excel/CSV はテキスト表として
そのまま Claude に渡す。OCR を挟まないぶんレイアウト崩れに強い。

金額は必ず文字列で受け取る。JSON の数値にすると float 経由で 1 円ずれる。
"""

from __future__ import annotations

import base64
import hashlib
import json
import mimetypes
import subprocess
import tempfile
from decimal import Decimal
from pathlib import Path
from typing import Any

import anthropic

from .config import Config
from .models import PurchaseOrder

# Claude API の document ブロック制限
MAX_REQUEST_BYTES = 30 * 1024 * 1024  # 仕様上 32MB。余裕を見て 30MB で止める

IMAGE_MEDIA_TYPES = {
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".gif": "image/gif",
    ".webp": "image/webp",
}
# iPhone 写真。そのままでは API が受け付けないので sips で JPEG 化する
CONVERTIBLE_IMAGE_SUFFIXES = {".heic", ".heif", ".tif", ".tiff", ".bmp"}

TAX_RATE_ENUM = ["10", "8", "0"]

PO_SCHEMA: dict[str, Any] = {
    "type": "object",
    "properties": {
        "po_number": {
            "type": "string",
            "description": "発注番号 / 注文番号。見つからなければ空文字。",
        },
        "order_date": {
            "type": "string",
            "description": "発注日。YYYY-MM-DD 形式。和暦は西暦へ直す。不明なら空文字。",
        },
        "delivery_date": {
            "type": "string",
            "description": "納期・納品日。YYYY-MM-DD 形式。不明なら空文字。",
        },
        "subject": {
            "type": "string",
            "description": "件名。なければ主要な品目名から簡潔に補う。",
        },
        "buyer": {
            "type": "object",
            "description": "発注元＝請求先。自社（受注側）ではないほうを取ること。",
            "properties": {
                "name": {"type": "string", "description": "会社名。株式会社等の法人格も含める。"},
                "department": {"type": "string", "description": "部署名。なければ空文字。"},
                "contact": {"type": "string", "description": "担当者名。なければ空文字。"},
                "postal_code": {"type": "string", "description": "郵便番号。なければ空文字。"},
                "address": {"type": "string", "description": "住所。なければ空文字。"},
            },
            "required": ["name", "department", "contact", "postal_code", "address"],
            "additionalProperties": False,
        },
        "items": {
            "type": "array",
            "description": "明細行。発注書に書かれた行をそのまま、順序を保って列挙する。",
            "items": {
                "type": "object",
                "properties": {
                    "name": {"type": "string", "description": "品名・作業内容。"},
                    "quantity": {
                        "type": "string",
                        "description": "数量。数字のみの文字列。記載がなければ 1。",
                    },
                    "unit": {"type": "string", "description": "単位（式・個・人日・時間 等）。不明なら 式。"},
                    "unit_price": {
                        "type": "string",
                        "description": "税抜の単価。数字のみの文字列（カンマ・円記号を含めない）。",
                    },
                    "tax_rate": {
                        "type": "string",
                        "enum": TAX_RATE_ENUM,
                        "description": "消費税率(%)。軽減税率対象は 8、不課税・非課税は 0、それ以外は 10。",
                    },
                    "note": {"type": "string", "description": "備考。なければ空文字。"},
                },
                "required": ["name", "quantity", "unit", "unit_price", "tax_rate", "note"],
                "additionalProperties": False,
            },
        },
        "payment_terms": {
            "type": "string",
            "description": "支払条件の記載（例: 月末締め翌月末払い）。なければ空文字。",
        },
        "notes": {"type": "string", "description": "備考欄の内容。なければ空文字。"},
        "stated_total": {
            "type": "string",
            "description": "発注書に印字されている合計金額（税込）。検算専用。なければ空文字。",
        },
        "withholding_tax": {
            "type": "boolean",
            "description": "源泉徴収の記載があるか。明示がなければ false。",
        },
        "confidence_notes": {
            "type": "array",
            "description": "読み取りに自信がない箇所・推測で埋めた箇所を日本語で列挙する。無ければ空配列。",
            "items": {"type": "string"},
        },
    },
    "required": [
        "po_number",
        "order_date",
        "delivery_date",
        "subject",
        "buyer",
        "items",
        "payment_terms",
        "notes",
        "stated_total",
        "withholding_tax",
        "confidence_notes",
    ],
    "additionalProperties": False,
}

SYSTEM_PROMPT = """\
あなたは日本の経理実務に精通した担当者です。渡された発注書（注文書・注文請書・見積承認書の\
場合もある）を読み、請求書を起こすために必要な項目を抽出します。

厳守事項:
- 書かれていないことを創作しない。読み取れない項目は空文字にし、confidence_notes に理由を書く。
- 金額はカンマ・円記号・全角数字を取り除いた半角数字の文字列で返す。小数が必要な場合のみ小数点を使う。
- unit_price は必ず税抜単価。発注書が税込単価しか示していない場合は税抜へ割り戻し、その旨を
  confidence_notes に記録する。
- 数量×単価が行の金額と合わない行があれば、発注書の金額欄を優先して単価を逆算し、
  confidence_notes に記録する。
- buyer は「発注する側（＝この請求書の宛先）」。発注書の中で「御中」が付く側や、発注者・注文者
  として署名している側を選ぶ。受注側（自社）の情報を buyer に入れてはいけない。
- 税率は発注書の記載に従う。「軽減税率対象」「※」等の印がある行は 8、対象外・不課税・非課税と
  読める行は 0、指定がなければ 10 とする。
- 値引き行やマイナス金額はそのまま負の unit_price として1行で表現する。
- 和暦（令和・平成）は西暦へ変換する。
"""


class ExtractionError(RuntimeError):
    pass


def sha256_of(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def _b64(data: bytes) -> str:
    # base64 に改行を含めると API 側で弾かれる
    return base64.standard_b64encode(data).decode("ascii")


def _convert_image(path: Path) -> tuple[bytes, str]:
    """HEIC / TIFF 等を macOS 標準の sips で JPEG に変換する。"""
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "converted.jpg"
        proc = subprocess.run(
            ["sips", "-s", "format", "jpeg", str(path), "--out", str(out)],
            capture_output=True,
            text=True,
        )
        if proc.returncode != 0 or not out.exists():
            raise ExtractionError(
                f"画像の変換に失敗しました（{path.name}）: {proc.stderr.strip() or 'sips エラー'}"
            )
        return out.read_bytes(), "image/jpeg"


def _tabular_to_text(path: Path) -> str:
    """Excel / CSV を Claude に読ませるためのテキスト表に落とす。"""
    suffix = path.suffix.lower()
    if suffix in (".csv", ".tsv", ".txt"):
        delimiter = "\t" if suffix == ".tsv" else ","
        import csv

        for encoding in ("utf-8-sig", "cp932", "utf-16"):
            try:
                with path.open(encoding=encoding, newline="") as fh:
                    rows = list(csv.reader(fh, delimiter=delimiter))
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ExtractionError(f"文字コードを判別できませんでした: {path.name}")
        return "\n".join(" | ".join(cell.strip() for cell in row) for row in rows)

    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    chunks: list[str] = []
    for sheet in wb.worksheets:
        lines = [f"### シート: {sheet.title}"]
        for row in sheet.iter_rows(values_only=True):
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            lines.append(" | ".join("" if c is None else str(c).strip() for c in row))
        if len(lines) > 1:
            chunks.append("\n".join(lines))
    wb.close()
    if not chunks:
        raise ExtractionError(f"中身が空のようです: {path.name}")
    return "\n\n".join(chunks)


def build_content(path: Path) -> list[dict[str, Any]]:
    """ファイル種別に応じて Claude へ渡す content ブロックを組み立てる。"""
    suffix = path.suffix.lower()
    size = path.stat().st_size
    if size > MAX_REQUEST_BYTES:
        raise ExtractionError(
            f"ファイルが大きすぎます（{size / 1024 / 1024:.1f}MB）。"
            f"30MB 以下に分割・圧縮してください: {path.name}"
        )

    instruction = {
        "type": "text",
        "text": "この発注書から請求書に必要な項目を抽出してください。",
    }

    if suffix == ".pdf":
        return [
            {
                "type": "document",
                "source": {
                    "type": "base64",
                    "media_type": "application/pdf",
                    "data": _b64(path.read_bytes()),
                },
            },
            instruction,
        ]

    if suffix in IMAGE_MEDIA_TYPES:
        return [
            {
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": IMAGE_MEDIA_TYPES[suffix],
                    "data": _b64(path.read_bytes()),
                },
            },
            instruction,
        ]

    if suffix in CONVERTIBLE_IMAGE_SUFFIXES:
        data, media_type = _convert_image(path)
        return [
            {
                "type": "image",
                "source": {"type": "base64", "media_type": media_type, "data": _b64(data)},
            },
            instruction,
        ]

    if suffix in (".xlsx", ".xlsm", ".csv", ".tsv", ".txt"):
        table = _tabular_to_text(path)
        return [
            {
                "type": "text",
                "text": f"以下は発注書ファイル `{path.name}` の内容です。\n\n```\n{table}\n```",
            },
            instruction,
        ]

    if suffix == ".xls":
        raise ExtractionError(
            f"旧形式の .xls には未対応です。Excel で .xlsx として保存し直してください: {path.name}"
        )

    guessed = mimetypes.guess_type(path.name)[0] or "不明"
    raise ExtractionError(f"未対応のファイル形式です（{guessed}）: {path.name}")


def _client(cfg: Config) -> anthropic.Anthropic:
    # 明示的なキーは渡さない。SDK が ANTHROPIC_API_KEY →
    # ANTHROPIC_AUTH_TOKEN → ant のプロファイル の順に解決する。
    del cfg
    return anthropic.Anthropic()


def extract(path: Path, cfg: Config) -> PurchaseOrder:
    """発注書ファイル 1 件を PurchaseOrder に変換する。"""
    content = build_content(path)
    client = _client(cfg)

    try:
        response = client.messages.create(
            model=cfg.api.model,
            max_tokens=cfg.api.max_tokens,
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": content}],
            output_config={
                "effort": cfg.api.effort,
                "format": {"type": "json_schema", "schema": PO_SCHEMA},
            },
        )
    except anthropic.AuthenticationError as exc:
        raise ExtractionError(
            "Claude API の認証に失敗しました。ANTHROPIC_API_KEY を設定してください。\n"
            "  キーの発行: https://console.anthropic.com/settings/keys\n"
            f"  詳細: {exc}"
        ) from exc
    except anthropic.NotFoundError as exc:
        raise ExtractionError(
            f"モデル ID が見つかりません: {cfg.api.model}\n"
            f"  config/company.toml の [api] model を確認してください。詳細: {exc}"
        ) from exc
    except anthropic.RateLimitError as exc:
        raise ExtractionError(f"レート制限に達しました。時間をおいて再実行してください: {exc}") from exc
    except anthropic.APIStatusError as exc:
        raise ExtractionError(f"Claude API がエラーを返しました（{exc.status_code}）: {exc}") from exc
    except anthropic.APIConnectionError as exc:
        raise ExtractionError(f"Claude API へ接続できませんでした: {exc}") from exc

    if response.stop_reason == "refusal":
        detail = getattr(response.stop_details, "explanation", None) or "理由の説明はありません"
        raise ExtractionError(f"Claude が処理を拒否しました: {detail}")
    if response.stop_reason == "max_tokens":
        raise ExtractionError(
            "出力が max_tokens に達して途中で切れました。"
            "config/company.toml の [api] max_tokens を増やしてください。"
        )

    text = next((b.text for b in response.content if b.type == "text"), None)
    if not text:
        raise ExtractionError("Claude から抽出結果が返りませんでした。")

    try:
        # parse_float=Decimal で、モデルが数値を返した場合も丸め誤差を持ち込まない
        data = json.loads(text, parse_float=Decimal)
    except json.JSONDecodeError as exc:
        raise ExtractionError(f"抽出結果を JSON として解釈できませんでした: {exc}") from exc

    data["source_file"] = str(path)
    data["source_sha256"] = sha256_of(path)
    data["extraction_warnings"] = data.pop("confidence_notes", [])

    order = PurchaseOrder.from_dict(data)
    if not order.items:
        raise ExtractionError(
            "明細を 1 行も読み取れませんでした。発注書のスキャン品質を確認するか、"
            "`seikyu new` で手入力してください。"
        )
    if not order.buyer.name:
        order.extraction_warnings.append("請求先の会社名を読み取れませんでした。確認画面で補ってください。")

    return order
